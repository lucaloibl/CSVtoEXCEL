import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import shutil
from openpyxl import load_workbook

# Global variable to store the selected background Excel file path.
background_excel_path = None

def excel_col_to_num(col_letters: str) -> int:
    """
    Convert Excel column letters (A, B, C, ... Z, AA, AB, ...) to a 1-based column index.
    For example, A -> 1, B -> 2, Z -> 26, AA -> 27, AB -> 28, etc.
    """
    col_letters = col_letters.upper().strip()
    result = 0
    for char in col_letters:
        if not ('A' <= char <= 'Z'):
            raise ValueError(f"Invalid column letter: {col_letters}")
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def update_sheet_dropdown():
    """Load the workbook and update the dropdown with available sheet names."""
    global target_sheet_var, background_excel_path
    try:
        wb = load_workbook(background_excel_path)
        sheet_names = wb.sheetnames
        if sheet_names:
            target_sheet_var.set(sheet_names[0])
            # Update the OptionMenu widget.
            menu = sheet_option["menu"]
            menu.delete(0, "end")
            for sheet in sheet_names:
                menu.add_command(label=sheet, command=lambda value=sheet: target_sheet_var.set(value))
        else:
            target_sheet_var.set("Sheet1")
    except Exception as e:
        messagebox.showerror("Error", f"Could not load sheet names from the background Excel file.\n{e}")

def select_background_file():
    """
    Opens a file dialog for the user to select the background Excel file.
    Stores the file path globally and updates the preview label and sheet dropdown.
    """
    global background_excel_path
    selected_file = filedialog.askopenfilename(
        title="Select Background Excel File",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if selected_file:
        background_excel_path = os.path.expanduser(selected_file)
        bg_label.config(text=f"Background file: {os.path.basename(background_excel_path)}")
        print("Background Excel file set to:", background_excel_path)
        messagebox.showinfo("Background File Selected", f"Background Excel file set to:\n{background_excel_path}")
        update_sheet_dropdown()
    else:
        messagebox.showwarning("No File Selected", "No background Excel file was selected.")

def preview_excel():
    """
    Opens a new window to display a preview of the target sheet in the background Excel file.
    The preview is shown in a Treeview widget which displays the first 10 rows in a tabular format.
    """
    global background_excel_path, target_sheet_var
    if not background_excel_path or not os.path.exists(background_excel_path):
        messagebox.showerror("Error", "No valid background Excel file set. Please select one first.")
        return
    try:
        wb = load_workbook(background_excel_path, data_only=True)
        sheet_name = target_sheet_var.get()
        if sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{sheet_name}' not found in the background Excel file.")
            return
        ws = wb[sheet_name]
        
        # Retrieve up to 10 rows from the sheet.
        preview_data = []
        max_rows = 10
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            preview_data.append([cell for cell in row])
            if i >= max_rows:
                break
        
        if not preview_data:
            messagebox.showinfo("Preview", "Sheet is empty.")
            return
        
        # Determine maximum columns in the preview.
        max_cols = max(len(row) for row in preview_data)
        col_names = [f"Col {i+1}" for i in range(max_cols)]
        
        # Create a new Toplevel window for the preview.
        preview_win = tk.Toplevel(root)
        preview_win.title(f"Preview of Sheet '{sheet_name}'")
        
        # Create a Treeview widget.
        tree = ttk.Treeview(preview_win, columns=col_names, show='headings')
        for col in col_names:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")
        
        # Insert rows into the Treeview.
        for row in preview_data:
            # Pad the row if it's shorter than max_cols.
            padded_row = list(row) + [''] * (max_cols - len(row))
            tree.insert("", "end", values=padded_row)
        
        tree.pack(expand=True, fill="both")
        
        # Add a vertical scrollbar.
        scrollbar = ttk.Scrollbar(preview_win, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        
    except Exception as e:
        messagebox.showerror("Error", f"Could not preview the Excel file.\n{e}")

def download_updated_excel():
    """Allows the user to download the updated background Excel file to a chosen location."""
    global background_excel_path
    if not background_excel_path or not os.path.exists(background_excel_path):
        messagebox.showerror("Error", "No valid background Excel file set.")
        return
    save_path = filedialog.asksaveasfilename(
        title="Download Updated Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if save_path:
        try:
            shutil.copy(background_excel_path, os.path.expanduser(save_path))
            messagebox.showinfo("Download Success", f"Updated Excel file downloaded to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not download Excel file.\n{e}")
    else:
        messagebox.showwarning("No Destination Selected", "No download destination was selected.")

def process_csv_file(file_path):
    """
    Processes the selected CSV file:
      - Reads the CSV.
      - Manipulates the data (sorts evaluations and adds missing ones with a count of 0).
      - Updates only the transaction count (the second column) into the specified region
        in the selected sheet of the background Excel file.
    """
    global background_excel_path, target_sheet_var
    if not background_excel_path or not os.path.exists(background_excel_path):
        messagebox.showerror("Error", "Background Excel file not found. Please set a valid background Excel file first.")
        return

    try:
        # Get and validate start row and start column from entries.
        row_str = start_row_entry.get().strip()
        col_str = start_col_entry.get().strip()
        
        try:
            start_row = int(row_str)
        except ValueError:
            raise ValueError(f"Invalid start row value: {row_str}. It must be an integer.")
        
        try:
            start_col = int(col_str)
        except ValueError:
            start_col = excel_col_to_num(col_str)
        
        # The target sheet is selected from the dropdown.
        target_sheet = target_sheet_var.get()

        # Expand user (~) and read the CSV file into a DataFrame.
        file_path = os.path.expanduser(file_path)
        df = pd.read_csv(file_path)

        print(f"Original rows count: {len(df)}")
        print(f"Column names: {df.columns.tolist()}")

        # Define the expected evaluations.
        expected_evals = [
             '{INCOME_CHILD_BENEFIT}', '{INCOME_NET}', '{INCOME_OTHER}', '{INCOME_PENSION}',
        '{INCOME_RENT}', '{INCOME_SUPPORT}', '{EXPENSE_CAR_INSURANCE}', '{EXPENSE_CONSTRUCTION_LOAN}',
        '{EXPENSE_HOUSE_SAVING}', '{EXPENSE_INSURANCE}', '{EXPENSE_INSURANCE_BU}', '{EXPENSE_INSURANCE_BUNDLE}',
        '{EXPENSE_INSURANCE_BUSINESS}', '{EXPENSE_INSURANCE_HAFTPFLICHT}', '{EXPENSE_INSURANCE_HAUSRAT}',
        '{EXPENSE_INSURANCE_HEALTH}', '{EXPENSE_INSURANCE_HEALTH_ADD}', '{EXPENSE_INSURANCE_LIFE}',
        '{EXPENSE_INSURANCE_RECHT}', '{EXPENSE_INSURANCE_RENTE}', '{EXPENSE_INSURANCE_UNFALL}',
        '{EXPENSE_INSURANCE_WOHNGEBAEUDE}', '{EXPENSE_LEASING}', '{EXPENSE_LOAN}', '{EXPENSE_LOAN_INTEREST}', '{EXPENSE_MICRO_LOAN}',
        '{EXPENSE_NON_LOAN_FINANCING}', '{EXPENSE_OTHER}', '{EXPENSE_OTHER_CHARGEBACK}', '{EXPENSE_PHONE}',
        '{EXPENSE_PROPERTY}', '{EXPENSE_PROPERTY_TAX}', '{EXPENSE_RENT}', '{EXPENSE_RENT_ADDITIONAL}',
        '{EXPENSE_SUPPORT}', '{CREDIT_BANK_MESSAGE}', '{CREDIT_CARRYOVER}', '{CREDIT_CASHDEPOSIT}',
        '{CREDIT_CHARGEBACK}', '{CREDIT_CHARGEBACK_CREDITCARD}', '{CREDIT_CHARGEBACK_INDIRECT}',
        '{CREDIT_CHARGEBACK_LOAN}', '{CREDIT_CHARGEBACK_OBJECTION}', '{CREDIT_CHARGEBACK_OBJECTION_CREDITCARD}',
        '{CREDIT_CHARGEBACK_OBJECTION_LOAN}', '{CREDIT_CHARGEBACK_OBJECTION_RENT}', '{CREDIT_CHARGEBACK_RENT}',
        '{CREDIT_CRYPTO}', '{CREDIT_DEPOSIT}', '{CREDIT_ECASH}', '{CREDIT_GAMBLING}', '{CREDIT_GAMBLING_LOTTO}',
        '{CREDIT_HEALTH_SUPPORT}', '{CREDIT_HOUSING_SUPPORT}', '{CREDIT_IDENT}', '{CREDIT_INKASSO}',
        '{CREDIT_INKASSO_MESSAGE}', '{CREDIT_INSURANCE}', '{CREDIT_INVOICE}', '{CREDIT_LEGAL}',
        '{CREDIT_LOAN}', '{CREDIT_MICRO_LOAN}', '{CREDIT_PARENTAL_SUPPORT}', '{CREDIT_PRIVATE_DRAWING}', '{CREDIT_PUBLIC_SECTOR}',
        '{CREDIT_REFUND}', '{CREDIT_RENT_ADDITIONAL}', '{CREDIT_SHOPPING}', '{CREDIT_STOCKTRADING}',
        '{CREDIT_TERMINAL}', '{CREDIT_UNEMPLOYMENT_SUPPORT}', '{CREDIT_VAT}', '{DEBIT_ATM}', '{DEBIT_CAR}',
        '{DEBIT_CAR_TAX}', '{DEBIT_CARRYOVER}', '{DEBIT_CHAMBER}', '{DEBIT_CHARGEBACK_FEE}', '{DEBIT_CREDITCARD}',
        '{DEBIT_CRYPTO}', '{DEBIT_CULTURE}', '{DEBIT_DEBITCARD}', '{DEBIT_DEPOSIT}', '{DEBIT_DIGITAL_SUBSCRIPTION}',
        '{DEBIT_DISPO}', '{DEBIT_DONATION}', '{DEBIT_ECASH}', '{DEBIT_EDUCATION}', '{DEBIT_EMPLOYEE_SALARY}',
        '{DEBIT_ENTERTAIN}', '{DEBIT_FEE}', '{DEBIT_FOOD_DRINK}', '{DEBIT_GAMBLING}', '{DEBIT_GAMBLING_LOTTO}',
        '{DEBIT_HEALTH}', '{DEBIT_INKASSO}', '{DEBIT_LEGAL}', '{DEBIT_LOAN_REPAY}', '{DEBIT_MEMBERSHIP_FEE}',
        '{DEBIT_OTHER_STANDING_ORDER}','{DEBIT_PUBLIC_SECTOR}', '{DEBIT_RESTAURANT}', '{DEBIT_SAVING}', '{DEBIT_SCHUFA}',
        '{DEBIT_SEIZURE}', '{DEBIT_SHOPPING}', '{DEBIT_STOCKTRADING}', '{DEBIT_TAX_BUSINESS}',
        '{DEBIT_TAX_INCOME}', '{DEBIT_TAX_VAT}', '{DEBIT_TRANSPORT}', '{DEBIT_TRAVEL}', '{DEBIT_VIDEOGAMES}',
        '{INFO_BANKINFORMATION}', '{INFO_DENIED_TRANSACTIONS}', '{INFO_OTHER}', '{INFO_RESERVATION}'
        ]

        # Assume the CSV file's first column is evaluations and the second column is the transaction count.
        eval_column = df.columns[0]
        count_column = df.columns[1]

        # Identify which evaluations are missing.
        existing_evals = df[eval_column].tolist()
        missing_evals = [eval_val for eval_val in expected_evals if eval_val not in existing_evals]

        # Add missing evaluations with transaction count = 0.
        for missing_eval in missing_evals:
            new_row = pd.DataFrame({eval_column: [missing_eval], count_column: [0]})
            df = pd.concat([df, new_row], ignore_index=True)
        print(f"Added {len(missing_evals)} missing rows with {count_column} = 0")

        # Create a sort order based on the expected evaluations.
        eval_order = {eval_val: i for i, eval_val in enumerate(expected_evals)}
        df['sort_order'] = df[eval_column].apply(lambda x: eval_order.get(x, 9999))
        df = df.sort_values('sort_order').drop('sort_order', axis=1)

        # Optionally, print any rows that still don't match expected evaluations.
        unmatched_rows = [row for row in df[eval_column] if row not in expected_evals]
        if unmatched_rows:
            print(f"Warning: {len(unmatched_rows)} rows didn't match any expected value:")
            for row in unmatched_rows[:5]:
                print(f"  - {row}")
            if len(unmatched_rows) > 5:
                print(f"  ... and {len(unmatched_rows)-5} more")

        # Open the background Excel file using openpyxl.
        wb = load_workbook(background_excel_path)
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
        else:
            ws = wb.create_sheet(target_sheet)

        # Write only the transaction count column into the worksheet starting at (start_row, start_col).
        # This assumes the CSV has exactly 2 columns and we want the second column.
        counts = df[count_column].tolist()
        for i, count in enumerate(counts, start=start_row):
            ws.cell(row=i, column=start_col, value=count)

        wb.save(background_excel_path)
        print(f"Updated sheet '{target_sheet}' starting at row {start_row}, col {start_col} in {background_excel_path}")
        messagebox.showinfo("Success", f"Background Excel file updated in sheet '{target_sheet}' at:\n{background_excel_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

def select_csv_file():
    """Opens a file dialog for CSV selection and processes the chosen file."""
    print("Select CSV button clicked; opening file dialog...")
    file_path = filedialog.askopenfilename(
        title="Select CSV File",
        filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
    )
    if file_path:
        print("CSV file selected:", file_path)
        process_csv_file(file_path)
    else:
        messagebox.showwarning("No File Selected", "No CSV file was selected.")

# --- GUI Setup ---

root = tk.Tk()
root.title("CSV to Excel Updater")
root.geometry("600x450")

# Create target_sheet_var after root is created.
target_sheet_var = tk.StringVar(root)
target_sheet_var.set("Sheet1")

# Frame for background file selection.
bg_frame = tk.Frame(root)
bg_frame.pack(pady=10)
bg_button = tk.Button(bg_frame, text="Set Background Excel File", command=select_background_file, font=("Helvetica", 12))
bg_button.pack(side=tk.LEFT, padx=5)
bg_label = tk.Label(bg_frame, text="Background file: Not set", font=("Helvetica", 10))
bg_label.pack(side=tk.LEFT, padx=5)

# Frame for sheet selection and preview.
sheet_frame = tk.Frame(root)
sheet_frame.pack(pady=10)
sheet_label = tk.Label(sheet_frame, text="Target Sheet:", font=("Helvetica", 12))
sheet_label.pack(side=tk.LEFT, padx=5)
sheet_option = tk.OptionMenu(sheet_frame, target_sheet_var, "Sheet1")
sheet_option.config(font=("Helvetica", 12))
sheet_option.pack(side=tk.LEFT, padx=5)
preview_button = tk.Button(sheet_frame, text="Preview Excel", command=preview_excel, font=("Helvetica", 12))
preview_button.pack(side=tk.LEFT, padx=5)

# Frame for specifying start row and column.
range_frame = tk.Frame(root)
range_frame.pack(pady=10)
start_row_label = tk.Label(range_frame, text="Start Row:", font=("Helvetica", 12))
start_row_label.pack(side=tk.LEFT, padx=5)
start_row_entry = tk.Entry(range_frame, font=("Helvetica", 12), width=5)
start_row_entry.insert(0, "2")
start_row_entry.pack(side=tk.LEFT, padx=5)
start_col_label = tk.Label(range_frame, text="Start Column (number or letters):", font=("Helvetica", 12))
start_col_label.pack(side=tk.LEFT, padx=5)
start_col_entry = tk.Entry(range_frame, font=("Helvetica", 12), width=10)
start_col_entry.insert(0, "1")  # Default can be "1" or "A"
start_col_entry.pack(side=tk.LEFT, padx=5)

# Button to select CSV file and update the Excel.
csv_button = tk.Button(root, text="Select CSV File & Update Excel", command=select_csv_file, font=("Helvetica", 12))
csv_button.pack(pady=20)

# Button to download the updated Excel.
download_button = tk.Button(root, text="Download Updated Excel", command=download_updated_excel, font=("Helvetica", 12))
download_button.pack(pady=10)

print("Starting main loop...")
root.mainloop()
print("Main loop terminated.")